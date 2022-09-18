# -*- coding: utf-8 -*-
"""
@Time ： 2022-09-02 21:15
@Auth ： 小胡
@File ：read_excl.py
@IDE ：PyCharm
@Motto：ABC(Always Be Coding)

"""
import openpyxl
import json
import time
r = '失败截图：' +time.strftime("%Y-%m-%d  %H:%M:%S") + '.png'
print(r)
def read_excel(file_name,sheet_name):
    book = openpyxl.load_workbook(file_name)
    #通过sheet名读取文件中的数据
    down_sheet = book[sheet_name]
    #获取最大行数
    rows = down_sheet.max_row
    #获取最大列数
    cols = down_sheet.max_column
    case_data = []  # 存放所有数据
    for i in range(2, rows + 1):
        case = {}
        for j in range(1, cols + 1):
            # 读取每个单元格的值
            value = down_sheet.cell(i, j).value
            # 将表头与内容形成键值对,如：{'姓名': '小明', '年龄': 24, '性别': '男', '国籍': '中国'}
            case[down_sheet.cell(1, j).value] = value
        case['test_data'] = json.loads(case['test_data'])
        case['exp'] = json.loads(case['exp'])
        # 将字典插入到列表：
        case_data.append(case)

    print(case_data)
    return case_data
if __name__ == '__main__':
    read_excel('../data/login.xlsx','login')
