# -*- coding: utf-8 -*-
"""
@Time ： 2022-09-02 21:15
@Auth ： 小胡
@File ：read_excl.py
@IDE ：PyCharm
@Motto：ABC(Always Be Coding)

"""
import inspect
import sys

import openpyxl
import json

def read_excel(file_name,sheet_name):
    # 案例：把对应的用例数读出来成为列表格式：
    # file_name = 'excel.xlsx'
    book = openpyxl.load_workbook(file_name)
    # sheet_name = book.sheetnames
    # print(sheet_name)


    down_sheet = book[sheet_name]

    # down_sheet = book[sheet_name]
        # 遍历整个文件,读取最大行数和最大列数
    rows = down_sheet.max_row
    cols = down_sheet.max_column
    case_data = []  # 存放所有数据
    for i in range(2, rows + 1):
        case = {}
        for j in range(1, cols + 1):
            # 读取每个单元格的值
            value = down_sheet.cell(i, j).value

            # 将表头与内容形成键值对,如：{'姓名': '小明', '年龄': 24, '性别': '男', '国籍': '中国'}
            case[down_sheet.cell(1, j).value] = value

        # 将字典插入到列表：
        case_data.append(case)

    # print(case_data)

    return case_data
if __name__ == '__main__':
    read_excel('../data/excel.xlsx','add_member')
